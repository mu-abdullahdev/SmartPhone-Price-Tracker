import os
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook,load_workbook
from datetime import date
filename="SmartPhone Price Tracker.xlsx"
if not os.path.exists(filename):
    wb=Workbook()
    sheet=wb.active
    sheet.title="Price Tracker"
    sheet.append(["SmartPhone Name", "Retail_Price", "Sale_Price","Date"])
else:
    wb=load_workbook(filename)
    sheet=wb.active
mobiles_list=["https://priceoye.pk/mobiles/samsung/samsung-galaxy-a07","https://priceoye.pk/mobiles/oppo/oppo-a6x","https://priceoye.pk/mobiles/itel/itel-super-s26-ultra","https://priceoye.pk/mobiles/honor/honor-x7c"]
session = requests.session()
today=date.today()
history={}
for row in sheet.iter_rows(min_row=2,values_only=True):
    product_name=row[0]
    product_price=row[2]
    history[product_name]=product_price
for mobile in mobiles_list:
    response= session.get(mobile)
    soup=BeautifulSoup(response.text,"html.parser")
    name=soup.find("span",class_="product-title-text").text
    retail_price=soup.find("span",class_="summary-price line-through stock-info").text
    retail_price=int(retail_price.replace("Rs","").replace(",","").strip())
    sale_price=soup.find("span",class_="summary-price text-black price-size-lg bold").text
    sale_price=int(sale_price.replace("Rs","").replace(",","").strip())
    if name in history:
        price_diff = history[name] - sale_price
        if price_diff>0:
            print(f"Alert! {name} price dropped by {price_diff} Rs!")
        elif price_diff<0:
            print(f"Alert! {name} price goes up by {price_diff} Rs!")
        else:
            print(f"The price is same!")
    sheet.append([name,retail_price,sale_price,today])
wb.save(filename)